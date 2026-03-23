from openpyxl import load_workbook

# 加载Excel文件
wb = load_workbook('Timesheet Report-20260316104428.xlsx')
sheet = wb.active

# 获取所有列名（第一行）
headers = []
for cell in sheet[1]:
    if cell.value:
        headers.append(cell.value)

print("Excel表的列名:")
for i, header in enumerate(headers, 1):
    print(f"{i}. {header}")

print(f"\n总共有 {len(headers)} 列")

# 显示前几行数据
print("\n前5行数据:")
for row_idx in range(1, min(6, sheet.max_row + 1)):
    row_data = []
    for col_idx in range(1, len(headers) + 1):
        cell = sheet.cell(row=row_idx, column=col_idx)
        row_data.append(cell.value)
    print(row_data)
